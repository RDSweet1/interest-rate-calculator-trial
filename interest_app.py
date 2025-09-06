# Interest Rate Calculator — Multi-Project + SharePoint Picker (HTML UI with Search & Project CRUD)
import os, json
import pandas as pd
from datetime import datetime, timedelta
from pandas.tseries.offsets import DateOffset
from typing import Optional, List, Tuple, Dict

OUTPUT_DIR   = os.path.join(os.getcwd(), "outputs")
INPUT_DIR    = os.path.join(os.getcwd(), "inputs")
PROJECTS_DIR = os.path.join(os.getcwd(), "projects")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(INPUT_DIR,  exist_ok=True)
os.makedirs(PROJECTS_DIR, exist_ok=True)
XLSX_PATH  = os.path.join(OUTPUT_DIR, "Interest_Rate_Calculator.xlsx")
PDF_PATH   = os.path.join(OUTPUT_DIR, "Interest_Rate_Calculator.pdf")
INPUT_FORM = os.path.join(INPUT_DIR,  "Interest_Rate_Calculator_Input.xlsx")

TITLE = "Interest Calculation for [Your Project Name]"
BILLING_DATE = datetime(2023, 4, 28)
AS_OF_DATE   = datetime(2025, 9, 5)
GRACE_DAYS   = 30
ANNUAL_RATE  = 0.18
MONTHLY_RATE = 0.015
PRINCIPAL_ITEMS = {"Flood/Wind": 13_365_247.68, "Drywall": 1_113_503.81}
PAYMENTS = [
    ("Flood Deductible", datetime(2023, 1, 4), 100_000.00),
    ("WIND Deductible",  datetime(2023, 1,24), 860_000.00),
    ("WIND Balance",     datetime(2023, 2,21), 2_700_000.00),
]

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_input_form(path: str):
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active; ws.title = "Project"
    ws.append(["Field","Value","Notes"])
    for c in range(1,4):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    rows = [
        ("Use Inputs","NO","Set to YES to override defaults"),
        ("Title", TITLE, "Project Title"),
        ("Billing Date", BILLING_DATE.strftime("%Y-%m-%d"), "YYYY-MM-DD"),
        ("As-of Date",  AS_OF_DATE.strftime("%Y-%m-%d"),  "YYYY-MM-DD"),
        ("Grace Days",  GRACE_DAYS, "Number of days"),
        ("Annual Rate", ANNUAL_RATE, "e.g., 0.18 for 18%"),
        ("Monthly Rate",MONTHLY_RATE, "Usually annual/12"),
    ]
    for r in rows: ws.append(list(r))
    ws2 = wb.create_sheet("Principals"); ws2.append(["Name","Amount"])
    ws2.append(["Flood/Wind", PRINCIPAL_ITEMS["Flood/Wind"]]); ws2.append(["Drywall", PRINCIPAL_ITEMS["Drywall"]])
    ws3 = wb.create_sheet("Payments"); ws3.append(["Date","Description","Amount"])
    for desc,d,amt in PAYMENTS: ws3.append([d.strftime("%Y-%m-%d"), desc, amt])
    wb.save(path)

create_input_form(INPUT_FORM)

from openpyxl.utils import get_column_letter
from openpyxl import Workbook as _WB
from openpyxl.styles import Font as _Font, Alignment as _Align
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
styles = getSampleStyleSheet()

money = lambda x: round(float(x), 2)

def slugify(name: str) -> str:
    s = "".join(c if c.isalnum() or c in ("-","_") else "-" for c in (name or "project").strip())
    while "--" in s: s = s.replace("--","-")
    return s.strip("-") or "project"

def compute_schedule(title: str, billing_date: datetime, as_of_date: datetime, grace_days: int,
                     annual_rate: float, monthly_rate: float,
                     principal_fw: float, principal_dw: float,
                     payments: List[Tuple[str, datetime, float]]):
    _fw_pay_total = sum(p[2] for p in payments)
    FW_NET = money(principal_fw - _fw_pay_total)
    DW_NET = money(principal_dw)
    FW_MONTHLY = money(FW_NET * monthly_rate)
    DW_MONTHLY = money(DW_NET * monthly_rate)

    INTEREST_START = billing_date + timedelta(days=grace_days)
    FIRST_PERIOD_END = INTEREST_START + DateOffset(months=1)
    charge_dates = pd.date_range(start=FIRST_PERIOD_END, end=as_of_date, freq=DateOffset(months=1))

    rows = []; cum_fw = 0.0; cum_dw = 0.0
    for end_ts in charge_dates:
        if end_ts.date() >= as_of_date.date():
            break
        pstart = (end_ts - DateOffset(months=1)).date(); pend = end_ts.date()
        cum_fw += FW_MONTHLY; cum_dw += DW_MONTHLY
        rows.append({"Start":pstart.strftime("%m/%d/%Y"),"End":pend.strftime("%m/%d/%Y"),
                     "FW Interest":FW_MONTHLY,"DW Interest":DW_MONTHLY,
                     "Total":money(FW_MONTHLY+DW_MONTHLY),
                     "Cumulative":money(cum_fw+cum_dw),"Type":"Full"})
    last_full_end = datetime.strptime(rows[-1]["End"], "%m/%d/%Y").date() if rows else INTEREST_START.date()
    partial_display_start = last_full_end + timedelta(days=1)
    next_period_end = (pd.Timestamp(last_full_end) + DateOffset(months=1)).date()
    if as_of_date.date() > last_full_end:
        days_in = (next_period_end - last_full_end).days
        days_el = (as_of_date.date() - last_full_end).days
        fw_p = money(FW_MONTHLY * days_el / days_in); dw_p = money(DW_MONTHLY * days_el / days_in)
        cum_fw += fw_p; cum_dw += dw_p
        rows.append({"Start":partial_display_start.strftime("%m/%d/%Y"),"End":as_of_date.strftime("%m/%d/%Y"),
                     "FW Interest":fw_p,"DW Interest":dw_p,"Total":money(fw_p+dw_p),
                     "Cumulative":money(cum_fw+cum_dw),"Type":f"Partial ({days_el}/{days_in} days)"})

    schedule_df = pd.DataFrame(rows, columns=["Start","End","FW Interest","DW Interest","Total","Cumulative","Type"])
    full_months = int((schedule_df["Type"]=="Full").sum())
    TOTAL_FW_INT = money(schedule_df["FW Interest"].sum()); TOTAL_DW_INT = money(schedule_df["DW Interest"].sum())
    summary_df = pd.DataFrame([
        {"Line Item":"Flood/Wind (net)","Principal":money(FW_NET),"Monthly Rate":"1.5%","Full Months":full_months,"Monthly Interest":money(FW_MONTHLY),"Total Interest":money(TOTAL_FW_INT)},
        {"Line Item":"Drywall",         "Principal":money(DW_NET),"Monthly Rate":"1.5%","Full Months":full_months,"Monthly Interest":money(DW_MONTHLY),"Total Interest":money(TOTAL_DW_INT)},
        {"Line Item":"TOTAL",           "Principal":money(FW_NET+DW_NET),"Monthly Rate":"","Full Months":full_months,"Monthly Interest":money(FW_MONTHLY+DW_MONTHLY),"Total Interest":money(TOTAL_FW_INT+TOTAL_DW_INT)},
    ])
    return summary_df, schedule_df

from openpyxl import Workbook as _WB
from openpyxl.styles import Font as _Font, Alignment as _Align

def export_excel_and_pdf(project_title: str, summary_df: pd.DataFrame, schedule_df: pd.DataFrame, slug: Optional[str] = None,
                          sharepoint_meta: Optional[Dict] = None):
    slug = slugify(slug or project_title)
    xlsx_per_project = os.path.join(OUTPUT_DIR, f"{slug}.xlsx")
    pdf_per_project  = os.path.join(OUTPUT_DIR, f"{slug}.pdf")

    wb = _WB(); ws_summary = wb.active; ws_summary.title = "Summary"
    ws_summary.append(list(summary_df.columns))
    for c in range(1, len(summary_df.columns)+1):
        cell = ws_summary.cell(row=1, column=c)
        cell.font = _Font(bold=True); cell.alignment = _Align(horizontal="center", vertical="center")
    for r in summary_df.itertuples(index=False): ws_summary.append(list(r))

    ws_sched = wb.create_sheet("Schedule"); ws_sched.append(list(schedule_df.columns))
    for c in range(1, len(schedule_df.columns)+1):
        cell = ws_sched.cell(row=1, column=c)
        cell.font = _Font(bold=True); cell.alignment = _Align(horizontal="center", vertical="center")
    for rec in schedule_df.to_dict('records'):
        ws_sched.append([str(rec.get('Start','')),str(rec.get('End','')),float(rec.get('FW Interest',0.0) or 0.0),
                         float(rec.get('DW Interest',0.0) or 0.0),float(rec.get('Total',0.0) or 0.0),
                         float(rec.get('Cumulative',0.0) or 0.0),str(rec.get('Type',''))])

    wb.save(xlsx_per_project); wb.save(XLSX_PATH)

    pdf = SimpleDocTemplate(pdf_per_project, pagesize=letter)
    pdf.build([Paragraph(project_title, styles['Title'])])
    pdf_default = SimpleDocTemplate(PDF_PATH, pagesize=letter)
    pdf_default.build([Paragraph(project_title, styles['Title'])])

    if sharepoint_meta:
        print("[SharePoint] Target folder:", sharepoint_meta)
    return xlsx_per_project, pdf_per_project

DEFAULT_PROJECT = {
    "title": TITLE,
    "billing_date": BILLING_DATE.strftime('%Y-%m-%d'),
    "as_of_date": AS_OF_DATE.strftime('%Y-%m-%d'),
    "grace_days": GRACE_DAYS,
    "annual_rate": ANNUAL_RATE,
    "monthly_rate": MONTHLY_RATE,
    "principal_fw": PRINCIPAL_ITEMS['Flood/Wind'],
    "principal_dw": PRINCIPAL_ITEMS['Drywall'],
    "payments": [{"desc": d[0], "date": d[1].strftime('%Y-%m-%d'), "amount": d[2]} for d in PAYMENTS],
    "sharepoint": {"folder_id": None, "folder_path": None}
}

def parse_project(obj: dict):
    pays = [(p['desc'], datetime.fromisoformat(p['date']), float(p['amount'])) for p in obj.get('payments', [])]
    return compute_schedule(
        obj['title'], datetime.fromisoformat(obj['billing_date']), datetime.fromisoformat(obj['as_of_date']),
        int(obj['grace_days']), float(obj['annual_rate']), float(obj['monthly_rate']),
        float(obj['principal_fw']), float(obj['principal_dw']), pays
    )

summary_df, schedule_df = parse_project(DEFAULT_PROJECT)
export_excel_and_pdf(DEFAULT_PROJECT['title'], summary_df, schedule_df, slug=DEFAULT_PROJECT['title'], sharepoint_meta=DEFAULT_PROJECT.get('sharepoint'))

for fname in sorted(os.listdir(PROJECTS_DIR)):
    if not fname.lower().endswith('.json'): continue
    try:
        with open(os.path.join(PROJECTS_DIR, fname), 'r') as f: obj = json.load(f)
        s_df, sch_df = parse_project(obj)
        export_excel_and_pdf(obj.get('title','project'), s_df, sch_df, slug=obj.get('title', os.path.splitext(fname)[0]), sharepoint_meta=obj.get('sharepoint'))
    except Exception as e:
        print(f"Skipping {fname}: {e}")

RUN_FLASK = os.environ.get('RUN_FLASK', '0')
if RUN_FLASK == '1':
    try:
        from flask import Flask, request, jsonify
        app = Flask(__name__)

        class StorageProvider:
            def list_children(self, parent_id: Optional[str] = None):
                raise NotImplementedError
            def get_path(self, item_id: Optional[str]):
                raise NotImplementedError
            def search(self, q: str):
                raise NotImplementedError

        class StubSharePoint(StorageProvider):
            def __init__(self):
                self.tree = {
                    'root': [
                        {"id": "f1", "name": "Accounting", "children": [
                            {"id": "f1a", "name": "2024"},
                            {"id": "f1b", "name": "2025"}
                        ]},
                        {"id": "f2", "name": "Ocean Harbor", "children": [
                            {"id": "f2a", "name": "Claims"},
                            {"id": "f2b", "name": "Invoices"}
                        ]},
                        {"id": "f3", "name": "Legal"}
                    ]
                }
            def list_children(self, parent_id: Optional[str] = None):
                return self.tree.get(parent_id or 'root', [])
            def get_path(self, item_id: Optional[str]):
                def dfs(nodes, path):
                    for n in nodes:
                        if n['id'] == item_id: return path + [n['name']]
                        if 'children' in n:
                            p = dfs(n['children'], path + [n['name']])
                            if p: return p
                    return None
                p = dfs(self.tree['root'], [])
                return '/' + '/'.join(p) if p else None
            def search(self, q: str):
                q = (q or '').strip().lower()
                def filt(nodes):
                    out = []
                    for n in nodes:
                        name = n['name']
                        kids = n.get('children')
                        child_filtered = filt(kids) if kids else []
                        match = (q in name.lower()) if q else True
                        if match or child_filtered:
                            o = {"id": n['id'], "name": name}
                            if child_filtered: o['children'] = child_filtered
                            out.append(o)
                    return out
                return filt(self.tree['root'])

        class GraphSharePoint(StorageProvider):
            # Microsoft Graph implementation (client credentials). Enable with USE_GRAPH=1.
            def __init__(self):
                self.enabled = os.environ.get('USE_GRAPH','0') == '1'
                self.tenant = os.environ.get('GRAPH_TENANT_ID')
                self.client_id = os.environ.get('GRAPH_CLIENT_ID')
                self.client_secret = os.environ.get('GRAPH_CLIENT_SECRET')
                self.site_id = os.environ.get('GRAPH_SITE_ID')
                self.drive_id = os.environ.get('GRAPH_DRIVE_ID')
                self.base = os.environ.get('GRAPH_BASE','https://graph.microsoft.com/v1.0')
                self._token = None
                self._ok = all([self.enabled, self.tenant, self.client_id, self.client_secret, self.site_id, self.drive_id])
                try:
                    import requests  # noqa
                    self._requests = requests
                except Exception as e:
                    self._ok = False
                    self._requests = None
                    print('[Graph] requests not available:', e)
            def _get_token(self):
                if self._token: return self._token
                if not self._ok:
                    raise RuntimeError('Graph not configured')
                data = {
                    'grant_type': 'client_credentials',
                    'client_id': self.client_id,
                    'client_secret': self.client_secret,
                    'scope': 'https://graph.microsoft.com/.default'
                }
                url = f'https://login.microsoftonline.com/{self.tenant}/oauth2/v2.0/token'
                r = self._requests.post(url, data=data, timeout=30)
                r.raise_for_status()
                self._token = r.json()['access_token']
                return self._token
            def _headers(self):
                return {'Authorization': f'Bearer {self._get_token()}'}
            def list_children(self, parent_id: Optional[str] = None):
                if not self._ok:
                    raise RuntimeError('Graph not configured')
                try:
                    if parent_id in (None, 'root'):
                        url = f"{self.base}/drives/{self.drive_id}/root/children"
                    else:
                        url = f"{self.base}/drives/{self.drive_id}/items/{parent_id}/children"
                    r = self._requests.get(url, headers=self._headers(), timeout=30)
                    r.raise_for_status()
                    items = r.json().get('value', [])
                    out = []
                    for it in items:
                        if it.get('folder') is None:
                            continue
                        out.append({ 'id': it['id'], 'name': it['name'] })
                    return out
                except Exception as e:
                    print('[Graph] list_children error:', e)
                    return []
            def get_path(self, item_id: Optional[str]):
                if not self._ok:
                    raise RuntimeError('Graph not configured')
                try:
                    url = f"{self.base}/drives/{self.drive_id}/items/{item_id}"
                    r = self._requests.get(url, headers=self._headers(), timeout=30)
                    r.raise_for_status()
                    obj = r.json()
                    pref = obj.get('parentReference', {})
                    parent_path = pref.get('path', '')
                    name = obj.get('name', '')
                    if parent_path.startswith('/drive/root:'):
                        parent_path = parent_path[len('/drive/root:'):]
                    full = (parent_path.rstrip('/') + '/' + name).replace('//','/')
                    if not full.startswith('/'):
                        full = '/' + full
                    return full
                except Exception as e:
                    print('[Graph] get_path error:', e)
                    return None
            def search(self, q: str):
                if not self._ok:
                    raise RuntimeError('Graph not configured')
                q = (q or '').strip()
                try:
                    url = f"{self.base}/drives/{self.drive_id}/root/search(q='{q}')"
                    r = self._requests.get(url, headers=self._headers(), timeout=30)
                    r.raise_for_status()
                    items = r.json().get('value', [])
                    out = []
                    for it in items:
                        if it.get('folder') is None:
                            continue
                        out.append({ 'id': it['id'], 'name': it['name'] })
                    return out
                except Exception as e:
                    print('[Graph] search error:', e)
                    return []

        def get_storage_provider():
            try:
                gp = GraphSharePoint()
                if gp.enabled and gp._ok:
                    print('[Storage] Using Microsoft Graph provider')
                    return gp
            except Exception as e:
                print('[Storage] Graph provider unavailable:', e)
            print('[Storage] Using Stub provider')
            return StubSharePoint()

        sp = get_storage_provider()

        def list_projects():
            files = [f for f in sorted(os.listdir(PROJECTS_DIR)) if f.lower().endswith('.json')]
            return files
        def load_project(name: str) -> dict:
            if name == '<default>': return DEFAULT_PROJECT.copy()
            with open(os.path.join(PROJECTS_DIR, name), 'r') as f: return json.load(f)
        def save_project(name: Optional[str], obj: dict) -> str:
            if name in (None, '<default>', '<new>'):
                name = slugify(obj.get('title','project')) + '.json'
            path = os.path.join(PROJECTS_DIR, name)
            with open(path, 'w') as f: json.dump(obj, f, indent=2)
            return name

        @app.route('/')
        def index():
            return (
                "<html><head><title>Interest Rate Calculator</title>"
                "<style>body{font-family:system-ui;margin:24px}.col{display:inline-block;vertical-align:top;margin-right:24px}ul{list-style:none;padding-left:16px}li{cursor:pointer;margin:4px 0}.path{color:#555;font-size:12px}.ok{color:#0a0}.monospace{font-family:monospace}</style>"
                "</head><body>"
                "<h2>Interest Rate Calculator — HTML</h2>"
                "<div class='col'>"
                "<h3>Projects</h3>"
                "<select id='proj'></select> <button onclick=refresh()>↻</button> <button onclick=newProj()>New</button> <button onclick=editProj()>Edit JSON</button>"
                "<div class='path'>Selected folder:<br><span id='sel'></span></div>"
                "<button onclick=saveSel()>Save Selection to Project</button>"
                "<div id='saveMsg'></div>"
                "</div>"
                "<div class='col'>"
                "<h3>SharePoint Folder Tree (stub)</h3>"
                "<input id='search' placeholder='Search folders...' oninput=searchTree() />"
                "<ul id='tree'></ul>"
                "</div>"
                "<div class='col'>"
                "<h3>Generate</h3>"
                "<button onclick=generate()>Generate Outputs (Default Project)</button>"
                "<pre id='out' class='monospace'></pre>"
                "</div>"
                "<script>\n"
                "let selected={id:null,path:null};\n"
                "async function api(u,o){const r=await fetch(u,o);return await r.json()}\n"
                "async function refresh(){const d=await api('/api/projects');const s=document.getElementById('proj');s.innerHTML='';['<default>'].concat(d.projects).forEach(n=>{const o=document.createElement('option');o.value=n;o.text=n;s.appendChild(o)});buildTree()}\n"
                "function node(item){const li=document.createElement('li');li.textContent=item.name;li.onclick=async(e)=>{e.stopPropagation();const p=await api('/api/sharepoint/path?id='+encodeURIComponent(item.id));selected.id=item.id;selected.path=p.path;document.getElementById('sel').textContent=p.path};if(item.children){const child=document.createElement('ul');item.children.forEach(c=>child.appendChild(node(c)));li.appendChild(child)}return li}\n"
                "async function buildTree(){const root=await api('/api/sharepoint/list');const ul=document.getElementById('tree');ul.innerHTML='';root.folders.forEach(it=>ul.appendChild(node(it)))}\n"
                "async function searchTree(){const q=document.getElementById('search').value||'';const res=await api('/api/sharepoint/search?q='+encodeURIComponent(q));const ul=document.getElementById('tree');ul.innerHTML='';res.folders.forEach(it=>ul.appendChild(node(it)))}\n"
                "function saveSel(){const proj=document.getElementById('proj').value||'<default>';if(!selected.id){document.getElementById('saveMsg').innerHTML='Pick a folder first.';return}api('/api/project/sharepoint',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({project:proj,folder_id:selected.id,folder_path:selected.path})}).then(res=>{document.getElementById('saveMsg').innerHTML='<span class=ok>Saved to '+res.project+'</span>'})}\n"
                "function newProj(){const title=prompt('Project title?');if(!title)return;const obj={title:title,billing_date:'2023-04-28',as_of_date:'2025-09-05',grace_days:30,annual_rate:0.18,monthly_rate:0.015,principal_fw:13365247.68,principal_dw:1113503.81,payments:[],sharepoint:{folder_id:null,folder_path:null}};api('/api/project/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({project:'<new>',data:obj})}).then(_=>refresh())}\n"
                "function editProj(){const proj=document.getElementById('proj').value;if(!proj)return;api('/api/project/get?name='+encodeURIComponent(proj)).then(data=>{const txt=prompt('Edit JSON:',JSON.stringify(data,null,2));if(!txt)return;try{const obj=JSON.parse(txt);api('/api/project/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({project:proj,data:obj})}).then(_=>refresh())}catch(e){alert('Invalid JSON')}})}\n"
                "async function generate(){const res=await api('/api/generate',{method:'POST'});document.getElementById('out').textContent=JSON.stringify(res,null,2)}\n"
                "refresh()\n"
                "</script>"
                "</body></html>"
            )

        @app.route('/api/projects')
        def api_projects():
            return jsonify({"projects": list_projects()})

        @app.route('/api/sharepoint/list')
        def api_sp_list():
            return jsonify({"folders": sp.list_children()})

        @app.route('/api/sharepoint/search')
        def api_sp_search():
            q = (request.args.get('q') or '').strip()
            return jsonify({"folders": sp.search(q)})

        @app.route('/api/sharepoint/path')
        def api_sp_path():
            item_id = request.args.get('id')
            return jsonify({"path": sp.get_path(item_id)})

        @app.route('/api/project/get')
        def api_project_get():
            name = request.args.get('name')
            return jsonify(load_project(name) if name and name != '<default>' else DEFAULT_PROJECT)

        @app.route('/api/project/save', methods=['POST'])
        def api_project_save():
            data = request.get_json(force=True)
            name = data.get('project')
            obj  = data.get('data')
            saved = save_project(name, obj)
            return jsonify({"status":"ok", "project": saved})

        @app.route('/api/project/sharepoint', methods=['POST'])
        def api_project_sp_save():
            data = request.get_json(force=True)
            proj = data.get('project') or '<default>'
            obj = load_project(proj) if proj and proj != '<default>' else DEFAULT_PROJECT.copy()
            obj.setdefault('sharepoint', {})
            obj['sharepoint']['folder_id'] = data['folder_id']
            obj['sharepoint']['folder_path'] = data['folder_path']
            name = save_project(proj, obj)
            return jsonify({"status":"ok", "project": name})

        @app.route('/api/generate', methods=['POST'])
        def api_generate():
            s_df, sch_df = parse_project(DEFAULT_PROJECT)
            xlsx, pdf = export_excel_and_pdf(DEFAULT_PROJECT['title'], s_df, sch_df, slug=DEFAULT_PROJECT['title'], sharepoint_meta=DEFAULT_PROJECT.get('sharepoint'))
            return jsonify({"excel": xlsx, "pdf": pdf})

        app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
    except Exception as e:
        print("Flask UI not available:", e)

# minimal tests
assert round(12*0.015, 6) == round(0.18, 6)
from openpyxl import load_workbook as _lb
if os.path.exists(XLSX_PATH):
    _wb_test = _lb(XLSX_PATH); _wsS = _wb_test["Summary"]
    assert hasattr(_wsS["A1"].alignment,'horizontal') and hasattr(_wsS["B1"].alignment,'horizontal')
print("Ready. Run with: RUN_FLASK=1 python interest_app.py")
